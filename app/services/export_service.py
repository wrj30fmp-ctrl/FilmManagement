"""
数据导出服务

支持 CSV 和 Excel 格式的数据导出。
CSV 使用 UTF-8 BOM 编码，确保中文在 Excel 中正常显示。
"""

import csv
import logging
from pathlib import Path
from datetime import datetime

from app.database.connection import get_db
from app.config import get_export_dir
from app.utils.date_utils import CHINA_TZ

logger = logging.getLogger(__name__)

EXPORT_TABLES = {
    "库存记录": "film_inventory",
    "拍摄记录": "film_rolls",
    "冲洗记录": "development_records",
    "扫描记录": "scan_records",
    "归档记录": "archive_records",
    "库存流水": "inventory_transactions",
    "状态历史": "film_roll_status_history",
}


class ExportService:
    """数据导出服务"""

    def __init__(self):
        self.db = get_db()
        self.export_dir = get_export_dir()

    def export_table_to_csv(self, table_name: str, output_path: str | Path | None = None) -> Path | None:
        """将指定表导出为 CSV 文件

        Args:
            table_name: 数据库表名
            output_path: 可选，指定输出路径；不指定则自动生成

        Returns:
            导出文件路径，失败返回 None
        """
        self.export_dir.mkdir(parents=True, exist_ok=True)

        if output_path is None:
            timestamp = datetime.now(CHINA_TZ).strftime("%Y%m%d_%H%M%S")
            output_path = self.export_dir / f"{table_name}_{timestamp}.csv"
        else:
            output_path = Path(output_path)

        try:
            with self.db.get_connection() as conn:
                cursor = conn.execute(f"SELECT * FROM {table_name};")
                rows = cursor.fetchall()
                if not rows:
                    logger.warning(f"表 {table_name} 无数据，导出空文件。")
                    # 至少获取列名
                    cursor = conn.execute(f"PRAGMA table_info({table_name});")
                    columns = [col[1] for col in cursor.fetchall()]
                else:
                    columns = rows[0].keys()

            # 使用 UTF-8 BOM 编码，确保 Excel 正确识别中文
            with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(columns)
                for row in rows:
                    writer.writerow([row[col] for col in columns])

            logger.info(f"CSV 导出完成: {output_path}")
            return output_path
        except Exception as e:
            logger.error(f"CSV 导出失败: {e}")
            return None

    def export_all_to_csv(self) -> dict[str, Path | None]:
        """导出所有表为 CSV 文件

        Returns:
            {表名: 导出文件路径} 字典，失败时为 None
        """
        timestamp = datetime.now(CHINA_TZ).strftime("%Y%m%d_%H%M%S")
        results = {}
        for display_name, table_name in EXPORT_TABLES.items():
            safe_name = display_name.replace(" ", "_")
            output = self.export_dir / f"{safe_name}_{timestamp}.csv"
            results[display_name] = self.export_table_to_csv(table_name, output)
        return results

    def export_to_excel(self, output_path: str | Path | None = None) -> Path | None:
        """将所有数据导出为一个 Excel 文件（多 Sheet）

        Args:
            output_path: Excel 文件路径，不指定则自动生成

        Returns:
            导出文件路径，失败返回 None
        """
        try:
            from openpyxl import Workbook
        except ImportError:
            logger.error("未安装 openpyxl 库，无法导出 Excel。请运行 pip install openpyxl")
            return None

        self.export_dir.mkdir(parents=True, exist_ok=True)

        if output_path is None:
            timestamp = datetime.now(CHINA_TZ).strftime("%Y%m%d_%H%M%S")
            output_path = self.export_dir / f"胶片管理数据_{timestamp}.xlsx"
        else:
            output_path = Path(output_path)

        try:
            wb = Workbook()
            # 移除默认 Sheet
            wb.remove(wb.active)

            for display_name, table_name in EXPORT_TABLES.items():
                with self.db.get_connection() as conn:
                    cursor = conn.execute(
                        f"SELECT * FROM {table_name} WHERE deleted_at IS NULL;"
                    )
                    rows = cursor.fetchall()

                ws = wb.create_sheet(title=display_name[:31])  # Excel Sheet 名限制 31 字符

                if rows:
                    columns = list(rows[0].keys())
                    ws.append(columns)
                    for row in rows:
                        ws.append([row[col] for col in columns])

            # 统计汇总 Sheet
            ws_summary = wb.create_sheet(title="统计汇总")
            ws_summary.append(["表名", "记录总数"])
            for display_name, table_name in EXPORT_TABLES.items():
                with self.db.get_connection() as conn:
                    cursor = conn.execute(
                        f"SELECT COUNT(*) FROM {table_name} WHERE deleted_at IS NULL;"
                    )
                    count = cursor.fetchone()[0]
                ws_summary.append([display_name, count])

            wb.save(output_path)
            logger.info(f"Excel 导出完成: {output_path}")
            return output_path
        except Exception as e:
            logger.error(f"Excel 导出失败: {e}")
            return None
